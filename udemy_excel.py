import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# CONFIGURAÇÃO — cole aqui o valor do cookie "access_token"
# Veja as instruções abaixo de como obtê-lo
# ============================================================
ACCESS_TOKEN = 'MEzerJM3wNBSEn7zagKGTPr0QZmn1DPEas4S0BmdBvA:U4fGcEooDUkqtvoXqR6VaKkgdvHtjX5ndmZv2H0vfJg'
OUTPUT_FILE  = 'udemy_cursos.xlsx'
# ============================================================


def fetch_all_courses(token: str) -> list[dict]:
    print('Buscando cursos na Udemy...')
    courses, page = [], 1

    while True:
        url = (
            'https://www.udemy.com/api-2.0/users/me/subscribed-courses/'
            '?fields[course]=title,completion_ratio,num_lectures,url'
            '&ordering=-last_accessed'
            f'&page={page}&page_size=100'
        )
        resp = requests.get(
            url,
            headers={
                'Authorization': f'Bearer {token}',
                'Accept':        'application/json, text/plain, */*',
                'Referer':       'https://www.udemy.com/home/my-courses/learning/',
                'User-Agent':    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'x-udemy-authorization': f'Bearer {token}',
            },
            timeout=15,
        )

        if resp.status_code == 401:
            print('\n❌ Token inválido ou expirado. Siga as instruções para obter um novo.')
            sys.exit(1)
        if resp.status_code != 200:
            print(f'\n❌ Erro na API: {resp.status_code}\n{resp.text[:400]}')
            sys.exit(1)

        data = resp.json()
        batch = data.get('results', [])
        courses.extend(batch)
        print(f'   Página {page}: {len(batch)} cursos')

        if not data.get('next'):
            break
        page += 1

    print(f'\n✅ {len(courses)} cursos encontrados.\n')
    return courses


# ── Estilos ──────────────────────────────────────────────────
DARK_BLUE  = '1C3553'
WHITE      = 'FFFFFF'
GREEN_BG   = 'C6EFCE';  GREEN_FG  = '276221'
YELLOW_BG  = 'FFEB9C';  YELLOW_FG = '9C6500'
RED_BG     = 'FFC7CE';  RED_FG    = '9C0006'
LIGHT_GREY = 'F2F2F2'
BORDER_CLR = 'CCCCCC'

thin   = Side(style='thin', color=BORDER_CLR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(color): return PatternFill('solid', fgColor=color)

def status_label(pct):
    return '✅ Concluído' if pct == 100 else ('🔴 Não iniciado' if pct == 0 else '🟡 Em andamento')

def progress_bar(pct, size=20):
    filled = round(pct / 100 * size)
    return '█' * filled + '░' * (size - filled)


def build_excel(courses: list[dict]) -> None:
    courses.sort(key=lambda c: c['completion_ratio'], reverse=True)

    wb = openpyxl.Workbook()

    # ── Aba Cursos ───────────────────────────────────────────
    ws = wb.active
    ws.title = 'Cursos'

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value     = f'📚 Meus Cursos Udemy — {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    c.font      = Font(bold=True, size=14, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['#', 'Título do Curso', 'Progresso', 'Barra', 'Aulas', 'Status', 'Data']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color=WHITE, size=11)
        c.fill      = fill(DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border
    ws.row_dimensions[2].height = 22

    today = datetime.now().strftime('%d/%m/%Y')

    for i, course in enumerate(courses):
        row  = i + 3
        pct  = round(course['completion_ratio'] * 100)
        bg   = LIGHT_GREY if i % 2 == 0 else WHITE

        row_vals = [
            i + 1,
            course.get('title', '—'),
            pct / 100,
            progress_bar(pct),
            course.get('num_lectures') or '—',
            status_label(pct),
            today,
        ]

        for col, val in enumerate(row_vals, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.fill      = fill(bg)
            c.border    = border
            c.alignment = Alignment(vertical='center')

        # Centraliza colunas pontuais
        for col in [1, 3, 5, 6, 7]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row, column=3).number_format = '0%'
        ws.cell(row=row, column=4).font = Font(name='Courier New', size=10)

        # Coloração por status
        if pct == 100:
            for col in [3, 6]:
                ws.cell(row=row, column=col).fill = fill(GREEN_BG)
                ws.cell(row=row, column=col).font = Font(color=GREEN_FG, bold=True)
        elif pct == 0:
            for col in [3, 6]:
                ws.cell(row=row, column=col).fill = fill(RED_BG)
                ws.cell(row=row, column=col).font = Font(color=RED_FG)
        else:
            for col in [3, 6]:
                ws.cell(row=row, column=col).fill = fill(YELLOW_BG)
                ws.cell(row=row, column=col).font = Font(color=YELLOW_FG)

        ws.row_dimensions[row].height = 20

    for col, width in enumerate([5, 55, 12, 24, 8, 18, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A3'

    # ── Aba Resumo ───────────────────────────────────────────
    ws2 = wb.create_sheet('Resumo')

    total     = len(courses)
    done      = sum(1 for c in courses if round(c['completion_ratio'] * 100) == 100)
    ongoing   = sum(1 for c in courses if 0 < round(c['completion_ratio'] * 100) < 100)
    not_start = sum(1 for c in courses if round(c['completion_ratio'] * 100) == 0)
    avg_pct   = sum(round(c['completion_ratio'] * 100) for c in courses) / total if total else 0

    ws2.merge_cells('A1:B1')
    c           = ws2['A1']
    c.value     = '📊 Resumo Geral'
    c.font      = Font(bold=True, size=14, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    items = [
        ('Total de cursos',   total,             DARK_BLUE,  WHITE),
        ('✅ Concluídos',     done,              GREEN_BG,   GREEN_FG),
        ('🟡 Em andamento',   ongoing,           YELLOW_BG,  YELLOW_FG),
        ('🔴 Não iniciados',  not_start,         RED_BG,     RED_FG),
        ('📈 Progresso médio', f'{avg_pct:.1f}%', 'EBF3FB',  '2E75B6'),
    ]

    for i, (label, value, bg, fg) in enumerate(items):
        row = i + 2
        for col, val in enumerate([label, value], 1):
            c           = ws2.cell(row=row, column=col, value=val)
            c.fill      = fill(bg)
            c.font      = Font(bold=True, color=fg, size=12)
            c.border    = border
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[row].height = 26

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    # Gráfico de pizza
    chart_rows = [('Categoria', 'Qtd'), ('Concluídos', done), ('Em andamento', ongoing), ('Não iniciados', not_start)]
    for i, (label, val) in enumerate(chart_rows):
        ws2.cell(row=i + 1, column=4, value=label)
        ws2.cell(row=i + 1, column=5, value=val)

    pie = PieChart()
    pie.title  = 'Distribuição dos Cursos'
    pie.style  = 10
    pie.width  = 14
    pie.height = 12
    pie.add_data(Reference(ws2, min_col=5, min_row=1, max_row=4), titles_from_data=True)
    pie.set_categories(Reference(ws2, min_col=4, min_row=2, max_row=4))
    ws2.add_chart(pie, 'D2')

    ws2.column_dimensions['D'].hidden = True
    ws2.column_dimensions['E'].hidden = True

    wb.save(OUTPUT_FILE)
    print(f'📄 Arquivo salvo: {OUTPUT_FILE}')


if __name__ == '__main__':
    if ACCESS_TOKEN == 'SEU_ACCESS_TOKEN_AQUI':
        print('⚠️  Preencha ACCESS_TOKEN no topo do arquivo antes de executar.')
        print('    Siga as instruções do README para obter o token.')
        sys.exit(1)

    courses = fetch_all_courses(ACCESS_TOKEN)
    build_excel(courses)
